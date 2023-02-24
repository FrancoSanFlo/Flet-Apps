# this is the main file where we handle user input data

# modules
from flet import *
from controls import return_control_reference, add_url_control_reference, return_url_control_reference
from database import Database
from openpyxl import *
from openpyxl.drawing.image import Image
from form_helper import return_date, return_new_quote, Cotizacion, Categoria, FormHelper, Cliente, return_new_client_code, return_existence
from validations import rut_validation, phone_validation

control_map = return_control_reference()
categories_list = []
total_value_ctz = []


def return_cotizacion(data_quote):
    n_cotizacion = data_quote[0]
    rut = data_quote[1]
    cliente = data_quote[2]
    solicitado_por = data_quote[3]
    fecha_solicitud = data_quote[4]
    folio = ''
    factoring = ''
    fecha_factura = ''
    descripcion = data_quote[5]
    estado = 'NO ENVIADA'
    ganada = 'NO'
    entregado = 'NO'
    facturado = 'NO'
    pagado = 'NO'
    neto = 0
    iva = 0

    # TODO: VALIDAR
    for category in categories_list:
        neto += int(category.subtotal)

    iva = (int(neto) * 0.19) + int(neto)
    
    
    cotizacion = Cotizacion(
        n_cotizacion,
        rut,
        cliente,
        solicitado_por,
        fecha_solicitud,
        folio,
        factoring,
        fecha_factura,
        descripcion,
        estado,
        ganada,
        entregado,
        facturado,
        pagado,
        neto,
        iva
    )
    return cotizacion

def save_into_excel(n_cotizacion, rut, cliente, solicitado_por, fecha_solicitud, descripcion, neto):

    # FOR DESKTOP - COT
    # wb = load_workbook('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx', data_only=True)

    # FOR NOTEBOOK - COT
    wb = load_workbook('C:\\Users\\franc\\OneDrive\\Escritorio\\ejemplo_cot.xlsx', data_only=True)
    
    ws = wb.active
    max_row = int(ws.max_row) + 1
    min_column = ws.min_column

    db = Database.ConnectToDatabase()
    record = Database.LastRecord(db)
    record = list(record)
    record.pop(0)

    for i in range(len(record)):
        if i == 8:
            if record[i] == 1:
                ws.cell(max_row,min_column).value = 'ENVIADA'
            else:
                ws.cell(max_row,min_column).value = 'NO ENVIADA'
        else:
            if record[i] == 1:
                ws.cell(max_row,min_column).value = 'SI'
            elif record[i] == 0:
                ws.cell(max_row,min_column).value = 'NO'
            else:
                ws.cell(max_row,min_column).value = record[i]
            
        min_column += 1

    min_column = ws.min_column

    # FOR DESKTOP - COT SAVE
    # wb.save('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx')

    # # FOR NOTEBOOK - COT SAVE
    wb.save('C:\\Users\\franc\\OneDrive\\Escritorio\\ejemplo_cot.xlsx')
    wb.close()

    """PARA GUARDAR EL FORMATO DE COTIZACIÓN"""
    # FOR DESKTOP - FORMAT
    # wb_format = load_workbook('C:\\Users\\franc\\Desktop\\COTIZACIONES\\FORMATO CTZ PERICLTDA.xlsx', data_only=True)

    # FOR NOTEBOOK - FORMAT
    wb_format = load_workbook('C:\\Users\\franc\\OneDrive\\Escritorio\\COTIZACIONES\\FORMATO CTZ PERICLTDA.xlsx', data_only=True) 

    ws_format = wb_format.active

    counter = 1
    in_row = 18
    # 1, 2, 6, 7, 8, 9
    for ctg in categories_list:
        ws_format.cell(row=in_row, column=1).value = counter
        ws_format.cell(row=in_row, column=2).value = ctg.nombre_categoria
        ws_format.cell(row=in_row, column=6).value = ctg.tipo_unidad
        ws_format.cell(row=in_row, column=7).value = ctg.cantidad
        ws_format.cell(row=in_row, column=8).value = ctg.valor_unitario
        ws_format.cell(row=in_row, column=9).value = ctg.subtotal
        in_row += 1
        counter += 1


    ws_format['B7'] = f'COTIZACIÓN N° {n_cotizacion}'
    ws_format['B16'] = descripcion
    ws_format['C9'] = cliente
    ws_format['C13'] = solicitado_por
    ws_format['H9'] = rut
    ws_format['H11'] = fecha_solicitud
    ws_format['I28'] = neto
    ws_format['I59'] = neto

    img = Image('assets/images/peric.png')
    ws_format.add_image(img, 'A1')

    # FOR DESKTOP - FORMAT SAVE
    # wb_format.save(f'C:\\Users\\franc\\Desktop\\COTIZACIONES\\INGRESADAS\\COTIZACION_{n_cotizacion}.xlsx')

    # FOR NOTEBOOK - FORMAT SAVE
    wb_format.save(f'C:\\Users\\franc\\OneDrive\\Escritorio\\COTIZACIONES\\INGRESADAS\\COTIZACION_{n_cotizacion}.xlsx')

    wb_format.close()

def update_into_excel(n_cotizacion):
    # FOR DESKTOP
    # wb = load_workbook('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx', data_only=True)

    # FOR NOTEBOOK
    wb = load_workbook('C:\\Users\\franc\\OneDrive\\Escritorio\\ejemplo_cot.xlsx', data_only=True)

    ws = wb.active
    dato_row_update = 0

    db = Database.ConnectToDatabase()
    record = Database.SearchByQuote(db, [n_cotizacion])
    
    # VALIDATION FOR SEARCH BY QUOTE 
    if record == None:
        print("NO EXISTE EL DATO INGRESADO", ValueError)
    else:
        for row in ws.iter_rows(min_row=ws.min_row, max_col=ws.min_column, max_row=ws.max_row):
            for cell in row:
                if cell.value == record[0]:
                    dato_row_update = cell.row
        # INSERTION OF DATA
        for i in range(ws.min_column, ws.max_column+1):
            ws.cell(row=dato_row_update, column=i).value = record[i-1]
        print("EXCEL ACTUALIZADO")

    # FOR DESKTOP
    # wb.save('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx')
    
    # FOR NOTEBOOK
    wb.save('C:\\Users\\franc\\OneDrive\\Escritorio\\ejemplo_cot.xlsx')
    wb.close()

# TODO: SEND MESSAGE TO SCREEN, VALIDATION APPFROMQUOTE      
def update_input_data(control_map_key):
    for key, value in control_map.items():
        if key == 'AppFormQuote' and control_map_key == 'AppFormQuote':
            n_cotizacion = value.controls[0].content.controls[0].controls[0].content.controls[1].value

            solicitud = value.controls[0].content.controls[0].controls[3].content.controls[1].value

            folio = value.controls[0].content.controls[1].controls[0].content.controls[1].value
            factoring = value.controls[0].content.controls[1].controls[1].content.controls[1].value
            fecha_factura = value.controls[0].content.controls[1].controls[2].content.controls[1].value

            descripcion = value.controls[0].content.controls[2].controls[0].content.controls[1].value

            estado = value.controls[0].content.controls[3].controls[0].content.controls[1].value
            ganada = value.controls[0].content.controls[3].controls[1].content.controls[1].value
            entregado = value.controls[0].content.controls[3].controls[2].content.controls[1].value
            facturado = value.controls[0].content.controls[3].controls[3].content.controls[1].value
            pagado = value.controls[0].content.controls[3].controls[4].content.controls[1].value

            # VALIDATION
            if n_cotizacion == '' or solicitud == '' or folio == '' or factoring == '' or fecha_factura == '' or descripcion == '':
                return False, "Faltan datos", icons.ERROR_ROUNDED
            else:
                db = Database.ConnectToDatabase()
                Database.UpdateDatabase(
                    db, (solicitud, descripcion, estado, ganada, entregado, facturado, pagado, folio, fecha_factura, factoring, n_cotizacion)
                )
                db.close()
                update_into_excel(n_cotizacion)
                return True, "Actualizado satisfactoriamente", icons.CHECK_CIRCLE_OUTLINE_ROUNDED
                
        if key == 'AppEditClientForm' and control_map_key == 'AppEditClientForm':
            codigo_cliente = value.controls[0].content.controls[0].controls[0].content.controls[1].value

            rut = value.controls[0].content.controls[0].controls[1].content.controls[1].value
            cliente = value.controls[0].content.controls[0].controls[2].content.controls[1].value
            fono = value.controls[0].content.controls[0].controls[3].content.controls[1].value
            direccion = value.controls[0].content.controls[0].controls[4].content.controls[1].value

            #VALIDATION
            if codigo_cliente == '' or rut == '' or cliente == '' or fono == '' or direccion == '':
                return False, "Campos vacíos", icons.ERROR_ROUNDED
            else:
                rut_boolean, validated_rut = rut_validation(rut)
                phone_boolean, validated_phone = phone_validation(fono)
                if rut_boolean:
                    if phone_boolean:
                        if return_existence(validated_rut):
                            return False, "RUT Existente", icons.ERROR_ROUNDED
                        else:
                            db = Database.ConnectToDatabase()
                            Database.UpdateClientsDatabase(
                                db, (validated_rut, cliente, validated_phone, direccion, codigo_cliente)
                            )
                            db.close()
                            return True, "Actualizado satisfactoriamente", icons.CHECK_CIRCLE_OUTLINE_ROUNDED
                    else:
                        return False, "Teléfono incorrecto", icons.ERROR_ROUNDED 
                else:
                    return False, "Formato de RUT incorrecto", icons.ERROR_ROUNDED

def clean_data_fields():
    for key, value in control_map.items():
        if key == 'AppRegisterForm':
            for user_input in value.controls[0].content.controls[0].controls[:]:
                if user_input.content.controls[0].value == 'Número Cotización':
                    user_input.content.controls[1].value = return_new_quote()
                    user_input.content.controls[1].update()
                elif user_input.content.controls[0].value == 'Fecha':
                    user_input.content.controls[1].value = return_date()
                    user_input.content.controls[1].update()
                else:
                    user_input.content.controls[1].value = ''
                    user_input.content.controls[1].update()

            value.controls[0].content.controls[1].controls[0].content.controls[1].value = ''
            value.controls[0].content.controls[1].controls[0].content.controls[1].update()

            for i in range(0, 5):
                if value.controls[0].content.controls[2].controls[i].content.controls[1].value == 'GL':
                    pass
                else:
                    value.controls[0].content.controls[2].controls[i].content.controls[1].value = ''
                    value.controls[0].content.controls[2].controls[i].content.controls[1].update()

            value.controls[0].content.controls[3].controls[0].content.controls[1].value = ''
            value.controls[0].content.controls[3].controls[0].content.controls[1].update()

        if key == "AppDataTable":
            value.controls[0].controls[0].rows.clear()
            value.controls[0].controls[0].update()

def get_input_data():
    data_cotizacion = []
    category_exist = False

    for key, value in control_map.items():
        if key == 'AppRegisterForm':
            for user_input in value.controls[0].content.controls[0].controls[:]:

                if user_input.content.controls[1].value == '' or user_input.content.controls[1].value == 'DRAGO PERIC':
                    pass
                else:
                    data_cotizacion.append(str(user_input.content.controls[1].value).upper())

            user_description = value.controls[0].content.controls[1].controls[0].content.controls[1].value
            if user_description != '':
                data_cotizacion.append(str(user_description).upper())

            for user_input in value.controls[0].content.controls[2].controls[:]:
                if user_input.content.controls[1].value != '' and user_input.content.controls[0].value != 'Tipo Unidad' :
                    category_exist = True
            
    if category_exist:
        return False, "Categoría sin ingresar", icons.ERROR_ROUNDED
    else:
        if len(data_cotizacion) < 6:
            return False, "Faltan datos", icons.ERROR_ROUNDED
        else:
            if len(categories_list) < 1:
                return False, "No hay categorías ingresadas", icons.ERROR_ROUNDED
            else:
                ctz = return_cotizacion(data_cotizacion)

                db = Database.ConnectToDatabase()
                Database.InsertDatabase(
                    db, 
                    (
                        ctz.n_cotizacion, 
                        ctz.cliente, 
                        ctz.rut, 
                        ctz.solicitado_por,
                        ctz.descripcion,
                        ctz.fecha_solicitud,
                        int(ctz.neto),
                        int(ctz.iva),
                        ctz.estado,
                        ctz.ganada,
                        ctz.entregado,
                        ctz.facturado,
                        ctz.pagado,
                        ctz.folio,
                        ctz.fecha_factura,
                        ctz.factoring
                    )
                )
                db.close()

                save_into_excel(ctz.n_cotizacion, ctz.rut, ctz.cliente, ctz.solicitado_por, ctz.fecha_solicitud, ctz.descripcion, ctz.neto)
                clean_data_fields()
                categories_list.clear()
                total_value_ctz.clear()
                return True, "Cotización registrada", icons.CHECK_CIRCLE_OUTLINE_ROUNDED

# TODO: SEND MESSAGE TO SCREEN, VALIDATION          
def fill_quotes(e):
    db = Database.ConnectToDatabase()
    records = Database.ReadDatabase(db)
    for key, value in control_map.items():
        if key == 'AppFormQuote':
            if len(records) == 0:
                print('THE DATABASE IS EMPTY')
            else:
                for user_input in value.controls[0].content.controls[0].controls[:]:
                    if user_input.content.controls[0].value == 'Número Cotización':
                        user_input.content.controls[1].options.clear()
                        for quote in Database.ReadDatabase(db)[::-1]:
                            user_input.content.controls[1].options.append(dropdown.Option(quote[0]))
                        user_input.content.controls[1].update()
                
                value.controls[0].content.controls[6].controls[1].controls[1].content.visible = False

def register_category(e):
    for key, value in control_map.items():
        if key == 'AppRegisterForm':
            data = DataRow(cells=[])

            name_category_value = value.controls[0].content.controls[2].controls[0].content.controls[1].value
            type_unit_value = value.controls[0].content.controls[2].controls[1].content.controls[1].value
            amount_value = value.controls[0].content.controls[2].controls[2].content.controls[1].value
            unit_value = value.controls[0].content.controls[2].controls[3].content.controls[1].value
            subtotal_value = value.controls[0].content.controls[2].controls[4].content.controls[1].value

            #TODO: ARREGLAR TOTAL
            if name_category_value == '' or type_unit_value == '' or amount_value == '' or unit_value == '' or subtotal_value == '':
                print('SOME FIELD IS EMPTY')
            else:
                try:
                    categoria = Categoria(str(name_category_value).upper(), type_unit_value, int(amount_value), int(unit_value), int(subtotal_value))
                    categories_list.append(categoria)

                    total_value = 0 
                    total_value_ctz.append(int(categoria.subtotal))

                    # TODO: ORDENAR
                    # RECORRE LISTA DE VALORES DE SUBTOTALES PARA LA COTIZACIÓN
                    for sub in total_value_ctz:
                        total_value += sub
                    value.controls[0].content.controls[3].controls[0].content.controls[1].value = total_value
                    value.controls[0].content.controls[3].controls[0].content.controls[1].update()

                    for i in range(0, 5):
                        data.cells.append(
                            DataCell(
                                FormHelper(str(value.controls[0].content.controls[2].controls[i].content.controls[1].value).upper())
                            )
                        )
                        if value.controls[0].content.controls[2].controls[i].content.controls[1].value == 'GL':
                            pass
                        else:
                            value.controls[0].content.controls[2].controls[i].content.controls[1].value = ''
                            value.controls[0].content.controls[2].controls[i].content.controls[1].update()
                except TypeError as e:
                    print(e)

        if key == "AppDataTable":
            if len(data.cells) == 0:
                pass
            else:
                value.controls[0].controls[0].rows.append(data)
                value.controls[0].controls[0].update()

def get_client_data():
    client_data = []
    for key, value in control_map.items():
        if key == 'AppClientForm':
            for user_input in value.controls[0].content.controls[0].controls[:]:
                if user_input.content.controls[1].value == '':
                    print("SOME FIELD IS EMPTY")
                else:
                    client_data.append(user_input.content.controls[1].value)
    
    if len(client_data) == 5:
        rut_boolean, validated_rut = rut_validation(client_data[1])
        phone_boolean, validated_phone = phone_validation(client_data[3])

        if rut_boolean:
            if phone_boolean:
                client = Cliente(client_data[0], validated_rut,  client_data[2], validated_phone, client_data[4])
                if return_existence(client.rut):
                    return False, "RUT Existente", icons.ERROR_ROUNDED
                else:
                    db = Database.ConnectToDatabase()
                    Database.InsertDatabaseClientes(db, (client.codigo_cliente, client.rut, client.cliente, client.fono, client.direccion))
                    client_data.clear()
                    # NOTE: CLEAN DATA CLIENT FIELDS
                    for key, value in control_map.items():
                        if key == 'AppClientForm':
                            for user_input in value.controls[0].content.controls[0].controls[:]:
                                if user_input.content.controls[0].value == 'Código cliente':
                                    user_input.content.controls[1].value = return_new_client_code()
                                    user_input.content.controls[1].update()
                                else:
                                    user_input.content.controls[1].value = ''
                                    user_input.content.controls[1].update()
                    return True, "Ingresado satisfactoriamente", icons.CHECK_CIRCLE_OUTLINE_ROUNDED
            else:
                return False, "Teléfono incorrecto", icons.ERROR_ROUNDED 
        else:
            return False, "Formato de rut incorrecto", icons.ERROR_ROUNDED
    else:
        return False, "Cliente no ingresado", icons.ERROR_ROUNDED

# TODO: SEND MESSAGE TO SCREEN, VALIDATION          
def filling_clients(control_map_key):
    db = Database.ConnectToDatabase()
    records = Database.ReadDatabaseClients(db)
    for key, value in control_map.items():
        if key == 'AppRegisterForm' and control_map_key == 'AppRegisterForm':
            if len(records) == 0:
                print("THE DATABASE IS EMPTY")
            else:
                code = value.controls[0].content.controls[5].controls[0].controls[1].content.controls[0].value
                if code == 'Código cliente':
                    value.controls[0].content.controls[5].controls[0].controls[1].content.controls[1].options.clear()
                    for code in Database.ReadDatabaseClients(db)[::-1]:
                        value.controls[0].content.controls[5].controls[0].controls[1].content.controls[1].options.append(dropdown.Option(code[0]))
                    value.controls[0].content.controls[5].controls[0].controls[1].content.controls[1].update()
        if key == 'AppEditClientForm' and control_map_key == 'AppEditClientForm':
            if len(records) == 0:
                print("THE DATABASE IS EMPTY")
            else:
                code = value.controls[0].content.controls[0].controls[0].content.controls[0].value
                if code == 'Código cliente':
                    value.controls[0].content.controls[0].controls[0].content.controls[1].options.clear()
                    for code in Database.ReadDatabaseClients(db)[::-1]:
                        value.controls[0].content.controls[0].controls[0].content.controls[1].options.append(dropdown.Option(code[0]))
                    value.controls[0].content.controls[0].controls[0].content.controls[1].update()


#URL CONFIGURATION
def add_url_data():
    for key, value in control_map.items():
            if key == 'AppSettingsForm':
                URL_cot_file = value.controls[0].content.controls[0].controls[0].content.controls[1].value
                URL_lay_file = value.controls[0].content.controls[0].controls[1].content.controls[1].value
                URL_cot_company_file = value.controls[0].content.controls[0].controls[2].content.controls[1].value

                if URL_cot_file == '' or URL_lay_file == '' or URL_cot_company_file == '':
                    pass
                else:
                    add_url_control_reference("url_cot_file", URL_cot_file)
                    add_url_control_reference("url_lay_file", URL_lay_file)
                    add_url_control_reference("url_cot_company_file", URL_cot_company_file)
                    control_url = return_url_control_reference()
                    for key, value in control_url.items():
                        print(key, value)

# REUSABLE DB FUNCTIONS
def fill_register_clients(e):
    filling_clients('AppRegisterForm')

def fill_edit_clients(e):
    filling_clients('AppEditClientForm')

def update_quote_input_data():
    boolean_value, value, icon_name = update_input_data('AppFormQuote')
    return boolean_value, value, icon_name

def update_client_input_data():
    boolean_value, value, icon_name = update_input_data('AppEditClientForm')
    return boolean_value, value, icon_name

# Buttons
def return_quotes_button():
    return Container(
        alignment=alignment.center,
        content=ElevatedButton(
            on_click=lambda e: fill_quotes(e),
            bgcolor='#007C91',
            color="white",
            content=Row(
                alignment=MainAxisAlignment.CENTER,
                controls=[
                    Icon(
                        name=icons.REFRESH_ROUNDED,
                        size=16
                    ),
                    Text(
                        "Cargar cotizaciones",
                        size=12,
                        weight="bold",
                    ),
                ],
            ),
            style=ButtonStyle(
                shape={
                    "": RoundedRectangleBorder(radius=6),
                },
                color={
                    "": "white",
                },
            ),
            height=42,
            width=170,
        ),
    )

def return_category_register_button():
    return Container(
        alignment=alignment.center,
        content=ElevatedButton(
            on_click=lambda e: register_category(e),
            bgcolor='#007C91',
            color="white",
            content=Row(
                alignment=MainAxisAlignment.CENTER,
                controls=[
                    Icon(
                        name=icons.CATEGORY_ROUNDED,
                        size=16
                    ),
                    Text(
                        "Registar categoría",
                        size=12,
                        weight="bold",
                    ),
                ],
            ),
            style=ButtonStyle(
                shape={
                    "": RoundedRectangleBorder(radius=6),
                },
                color={
                    "": "white",
                },
            ),
            height=42,
            width=170,
        ),
    )

def return_clients_button():
    return Container(
        alignment=alignment.center,
        content=ElevatedButton(
            on_click=lambda e: fill_register_clients(e),
            bgcolor='#007C91',
            color="white",
            content=Row(
                alignment=MainAxisAlignment.CENTER,
                controls=[
                    Icon(
                        name=icons.REFRESH_ROUNDED,
                        size=16
                    ),
                    Text(
                        "Cargar clientes",
                        size=12,
                        weight="bold",
                    ),
                ],
            ),
            style=ButtonStyle(
                shape={
                    "": RoundedRectangleBorder(radius=6),
                },
                color={
                    "": "white",
                },
            ),
            height=42,
            width=170,
        ),
    )

def return_edit_clients_button():
    return Container(
        alignment=alignment.center,
        content=ElevatedButton(
            on_click=lambda e: fill_edit_clients(e),
            bgcolor='#007C91',
            color="white",
            content=Row(
                alignment=MainAxisAlignment.CENTER,
                controls=[
                    Icon(
                        name=icons.REFRESH_ROUNDED,
                        size=16
                    ),
                    Text(
                        "Cargar clientes",
                        size=12,
                        weight="bold",
                    ),
                ],
            ),
            style=ButtonStyle(
                shape={
                    "": RoundedRectangleBorder(radius=6),
                },
                color={
                    "": "white",
                },
            ),
            height=42,
            width=170,
        ),
    )
