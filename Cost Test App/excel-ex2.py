import openpyxl
from openpyxl import *


wb = load_workbook('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx', data_only=True)
ws = wb.active
# max_row = ws.max_row

# for row in range(1, 101):
#     for col in list("ABCDEFGHIJKL"):
#         value = random.random()
#         idx = f"{col}{row}"
#         ws[idx] = value

# for i in range(ws.min_row, ws.max_column+1):
#     print(i)

# print(int(ws.cell(max_row,1).value) + 1)
# print("")
# print("\nMinimum row: {0}".format(ws.min_row))
# print("Maximum row: {0}".format(ws.max_row))
# print("Minimum column: {0}".format(ws.min_column))
# print("Maximum column: {0}".format(ws.max_column))

wb.close()
        
# wb.save('C:\\Users\\franc\\Desktop\\ejemplo_cot.xlsx')