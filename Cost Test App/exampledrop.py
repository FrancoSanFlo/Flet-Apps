from openpyxl import *
from openpyxl.drawing.image import Image

# wb = load_workbook('C:\\Users\\franc\\Desktop\\FORMATO CTZ PERICLTDA.xlsx', data_only=True)
# wb = load_workbook('C:\\Users\\franc\\OneDrive\\Escritorio\\ejemplo_cot.xlsx', data_only=True)
# ws = wb.active

# print(len(ws.cell(row=28, column=7).value))

# for row in ws.iter_rows(min_row=18, max_col=9, max_row=ws.max_row):
#     for cell in row:
#         if cell.value == "TOTAL NETO (CLP)":
#             print("ENCONTRADO NETO", cell.row, cell.column)
#         if cell.value == "PRECIO UNITARIO NETO DE LA PARTIDA CON GASTOS GENERALES Y UTILIDADES (CLP)":
#             print("ENCONTRADO EL VALOR TOTAL", cell.row, cell.column)

# ws['A18'] = 'HOLA'
# ws['B18'] = "HOLA"
# ws['F18'] = "HOLA"
# ws['G18'] = "HOLA"
# ws['H18'] = "HOLA"
# ws['I18'] = "HOLA"
# ws.cell(row=18, column=1).value = 'DESDE COORDENADA'

# ws['A19'] = 'HOLANDA'
# ws['B19'] = "HOLANDA"
# ws['F19'] = "HOLANDA"
# ws['G19'] = "HOLANDA"
# ws['H19'] = "HOLANDA"
# ws['I19'] = "HOLANDA"

# cotizacion = 1234

# img = Image('images/peric.png')
# ws.add_image(img, 'A1')

# wb.save('C:\\Users\\franc\\Desktop\\FORMATO CTZ PERICLTDA EDITADO.xlsx')
# wb.save(f'C:\\Users\\franc\\Desktop\\Cotizacion_{cotizacion}.xlsx')
# wb.close()

from validations import rut_validation, phone_validation
rut = '123456789'
rut2 = '12345678'
rut3 = '12345678-9'
rut4 = '1234567-8'

def return_rut(rut):
    if '-' in rut:
        if len(rut) == 9:
            return '{0}.{1}.{2}'.format(rut[:1], rut[1:4], rut[4:])
        else:
            return '{0}.{1}.{2}'.format(rut[:2], rut[2:5], rut[5:])
    else:
        return '{0}.{1}.{2}-{3}'.format(rut[:2], rut[2:5], rut[5:8], rut[-1])

# print(return_rut(rut))
# print(return_rut(rut1))
# print(return_rut(rut2))

# print(rut_validation(rut))
# print(rut_validation(rut2))
# print(rut_validation(rut4))
# print(rut_validation(rut3))
print(phone_validation('987654321'))
